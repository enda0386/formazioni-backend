import io
from datetime import date
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from database import db
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


@router.get("/riepilogo")
def riepilogo():
    with db() as con:
        cantieri = con.execute("SELECT * FROM v_riepilogo_cantieri").fetchall()
        tot_dip = con.execute("SELECT COUNT(*) FROM dipendenti WHERE attivo=1").fetchone()[0]
        scad = con.execute(
            "SELECT COUNT(*) FROM v_attestati WHERE giorni_alla_scadenza < 0"
        ).fetchone()[0]
        in_scad = con.execute(
            "SELECT COUNT(*) FROM v_attestati WHERE giorni_alla_scadenza BETWEEN 0 AND 60"
        ).fetchone()[0]
        visite_scad = con.execute(
            "SELECT COUNT(*) FROM v_visite WHERE giorni_alla_scadenza < 0"
        ).fetchone()[0]
        visite_in_scad = con.execute(
            "SELECT COUNT(*) FROM v_visite WHERE giorni_alla_scadenza BETWEEN 0 AND 60"
        ).fetchone()[0]
    return {
        "totale_dipendenti_attivi": tot_dip,
        "attestati_scaduti": scad,
        "attestati_in_scadenza_60gg": in_scad,
        "visite_scadute": visite_scad,
        "visite_in_scadenza_60gg": visite_in_scad,
        "per_cantiere": [dict(r) for r in cantieri],
    }


@router.get("/export/excel")
def export_excel():
    with db() as con:
        att = con.execute(
            "SELECT * FROM v_attestati ORDER BY cantiere, dipendente"
        ).fetchall()
        visite = con.execute(
            "SELECT * FROM v_visite ORDER BY cantiere, dipendente"
        ).fetchall()

    wb = Workbook()
    _sheet_attestati(wb, att)
    _sheet_visite(wb, visite)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"formazioni_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── helpers ──────────────────────────────────────────────────────────────────

_HDR = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_CTR = Alignment(horizontal="center", vertical="center")

def _fill_for(giorni):
    if giorni is None:
        return PatternFill("solid", fgColor="F0F0F0")
    if giorni < 0:
        return PatternFill("solid", fgColor="FFD0D0")
    if giorni <= 60:
        return PatternFill("solid", fgColor="FFF2CC")
    return PatternFill("solid", fgColor="D4EDDA")


def _sheet_attestati(wb: Workbook, rows):
    ws = wb.active
    ws.title = "Attestati"
    hdrs = ["Cantiere", "Dipendente", "Agenzia", "Corso", "Categoria",
            "Data Esecuzione", "Data Scadenza", "Giorni", "Stato", "Ente formatore"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HDR; c.fill = _HDR_FILL; c.alignment = _CTR
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        giorni = row["giorni_alla_scadenza"]
        fill = _fill_for(giorni)
        vals = [row["cantiere"], row["dipendente"], row["agenzia"],
                row["corso"], row["categoria"], row["data_esecuzione"],
                row["data_scadenza"], giorni, row["stato"], row["ente_formatore"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=9)
            c.fill = fill
            c.alignment = _CTR if ci > 5 else Alignment(horizontal="left", vertical="center")
    for ci, w in enumerate([22, 26, 16, 34, 14, 14, 14, 10, 12, 20], 1):
        ws.column_dimensions[chr(64+ci)].width = w
    ws.freeze_panes = "A2"


def _sheet_visite(wb: Workbook, rows):
    ws = wb.create_sheet("Visite Mediche")
    hdrs = ["Cantiere", "Dipendente", "Agenzia", "Tipo", "Data Visita",
            "Data Scadenza", "Giorni", "Esito", "Medico"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HDR
        c.fill = PatternFill("solid", fgColor="375623")
        c.alignment = _CTR
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        giorni = row["giorni_alla_scadenza"]
        fill = _fill_for(giorni)
        vals = [row["cantiere"], row["dipendente"], row["agenzia"],
                row["tipo"], row["data_visita"], row["data_scadenza"],
                giorni, row["esito"], row["medico"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=9)
            c.fill = fill
            c.alignment = _CTR if ci > 4 else Alignment(horizontal="left", vertical="center")
    for ci, w in enumerate([22, 26, 16, 14, 14, 14, 10, 18, 20], 1):
        ws.column_dimensions[chr(64+ci)].width = w
    ws.freeze_panes = "A2"
